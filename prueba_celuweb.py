from openpyxl import Workbook

from os import remove


from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from time import sleep
import pickle
from deep_translator import GoogleTranslator


### inicializacion del webdriver

service = Service(executable_path=r'c:\webdriver\msedgedriver.exe')
driver = webdriver.Edge(service=service)

driver.get("https://www.instagram.com")

sleep(1)
########
### inyeccion de cookies
m_cookies= pickle.load(open("cookies.pkl", "rb"))

for cookie in m_cookies:
    cookie["domain"]= ".instagram.com"
    try:
        driver.add_cookie(cookie)
    except:
        pass

driver.refresh()



driver.get("https://www.instagram.com/samsung/")
#sleep(1)

#encapsulado ul de la info superior (seguidores, seguidos, publicaciones)
wait = WebDriverWait(driver, 3)

all_info = wait.until(EC.presence_of_element_located((
    By.CSS_SELECTOR, 
    "ul.x78zum5.x1q0g3np.xieb3on"
)))

#lista de 3 elementos
all = all_info.find_elements(By.CSS_SELECTOR, "li")

publicaciones = all[0].find_element(By.CSS_SELECTOR, "span").text
seguidores = all[1].find_element(By.CSS_SELECTOR, "span").get_attribute("title")
seguidos = all[2].find_element(By.CSS_SELECTOR, "span").text

print(f"Publicaciones: {publicaciones}")
print(f"Seguidores: {seguidores}")
print(f"Seguidos: {seguidos}")



#descripcion
aux_descripcion = wait.until(EC.presence_of_element_located((
    By.CSS_SELECTOR, 
    "span._ap3a._aaco._aacu._aacx._aad7._aade"
)))
descripcion = aux_descripcion.text
desc_es = GoogleTranslator(source='en', target='es').translate(descripcion)




#nombre
aux_nombre = wait.until(EC.presence_of_element_located((
        By.CSS_SELECTOR, 
        "span.x1lliihq.x193iq5w.x6ikm8r.x10wlt62.xlyipyv.xuxw1ft"
    )))
nombre = aux_nombre.text
print(f"nombre: {nombre}")

#creacion archivo excel
wb= Workbook()
ws = wb.active
ws.title="Samgung"
ws.append(["nombre_cuenta", "numero_publicaciones", "numero_seguidores", "numero_seguidos", "descripcion"])

ws.cell(row=2, column=1, value=nombre)
ws.cell(row=2, column=2, value=publicaciones)
ws.cell(row=2, column=3, value=seguidores)  
ws.cell(row=2, column=4, value=seguidos)
ws.cell(row=2, column=5, value=desc_es)

    
wb.save("cuentas.xlsx")


